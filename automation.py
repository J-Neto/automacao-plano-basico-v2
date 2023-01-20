from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
import time
import os
import openpyxl
import pandas as pd


def OpenBrowser():

    # Using ChromeOptions() to supress the "bluetooth_adapter_winrt.cc" error
    # We aren't using bluetooth, so its fine to supress this error message
    options = webdriver.ChromeOptions()
    options.add_experimental_option("excludeSwitches", ["enable-logging"])

    # Looking for chromedriver
    driver = webdriver.Chrome(options=options, service=Service(
        ChromeDriverManager().install()))

    # Opening Chrome Browser
    driver.get("http://sistemas.anatel.gov.br/se/public/view/b/srd.php")
    time.sleep(5)

    return driver


def RegistersPerPage(driver, n):

    # Getting the number of registers field
    qtd_field = driver.find_element(By.XPATH, '//*[@id="rpp"]')

    # Perfoming a double click
    action = ActionChains(driver)
    action.double_click(qtd_field).perform()

    # Defining the number of registers and pressing ENTER (RETURN button)
    qtd_field.send_keys('250' + Keys.RETURN)

    time.sleep(3)


def Filter(driver, service_name):

    # Clicking on "Filtrar" element
    driver.find_element(By.XPATH, '//*[@id="tblFilter"]/span[5]').click()
    time.sleep(3)

    # Type the "Serviço" search input, searching for the service name and pressing ENTER (RETURN button)
    driver.find_element(
        By.XPATH, '//*[@id="fc_6"]').send_keys(service_name + Keys.RETURN)
    time.sleep(3)


def ChangeWorksheetName(wb, filepath):

    # Getting the worksheet
    ws = wb[wb.sheetnames[0]]

    # Changing the worksheet name
    ws.title = "Sheet1"
    wb.save(filepath)


def OpenExcel():

    # Setting env variable
    os.environ['FILEPATH'] = "C:/Mirante/Projects/automacao-plano-basico-v2/plano-basico.xlsx"
    filepath = os.getenv('FILEPATH')

    # Creating the workbook
    wb = openpyxl.Workbook()

    # Saving the workbook
    wb.save(filepath)

    # Loading workbook
    wb = openpyxl.load_workbook(
        r"C:/Mirante/Projects/automacao-plano-basico-v2/plano-basico.xlsx")

    ChangeWorksheetName(wb, filepath)


def GetDataToTable(driver):

    # Getting Table
    tbl = driver.find_element(
        By.XPATH, '//*[@id="aplTbl"]').get_attribute('outerHTML')

    # Converting to pandas dataframe
    df = pd.read_html(tbl)
    df = df[0]


def Search(driver):

    # Define 250 as the number os registers per page
    RegistersPerPage(driver, 250)

    # Filtering the service name as "TV"
    Filter(driver, "TV")

    GetDataToTable(driver)

    # Minimizing chrome browser
    driver.minimize_window()


def Close(driver):

    # Wait 5 seconds than close the Browser
    time.sleep(5)
    driver.close()


chrome = OpenBrowser()
Search(chrome)
# OpenExcel()
# Close(chrome)
